"""
Excel automation tools — read/write .xlsx files via openpyxl.

Windows only: uses local filesystem paths. No COM dependency.
"""

from __future__ import annotations

import json
import logging
import os
from pathlib import Path

from langchain.tools import tool

logger = logging.getLogger(__name__)


def _resolve_path(filepath: str) -> str:
    """Normalize and validate a file path."""
    path = Path(os.path.normpath(filepath)).resolve()
    return str(path)


def _check_openpyxl_installed() -> str | None:
    """Return an error message if openpyxl is not available, else None."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return "openpyxl is not installed. Run: pip install openpyxl"
    return None


@tool("read_excel_range", parse_docstring=True)
def read_excel_range_tool(filepath: str, sheet_name: str | None = None) -> str:
    """Read data from an Excel (.xlsx) workbook.

    Returns all data from the specified sheet (or first sheet by default)
    as a JSON array of rows.

    Args:
        filepath: Absolute or relative path to the .xlsx file to read.
        sheet_name: Name of the sheet to read. If omitted, reads the active/first sheet.
    """
    err = _check_openpyxl_installed()
    if err:
        return json.dumps({"error": err}, ensure_ascii=False)

    path = _resolve_path(filepath)
    if not os.path.isfile(path):
        return json.dumps({"error": f"File not found: {path}"}, ensure_ascii=False)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return json.dumps({"error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"}, ensure_ascii=False)
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else None for cell in row])

        wb.close()

        return json.dumps(
            {
                "filepath": path,
                "sheet_name": sheet_name,
                "total_rows": len(rows),
                "total_columns": len(rows[0]) if rows else 0,
                "data": rows,
            },
            indent=2,
            ensure_ascii=False,
        )
    except Exception as exc:
        logger.exception("Failed to read Excel file: %s", exc)
        return json.dumps({"error": f"Failed to read Excel file: {exc}"}, ensure_ascii=False)


@tool("write_excel_range", parse_docstring=True)
def write_excel_range_tool(filepath: str, sheet_name: str | None = None, data: str = "[]") -> str:
    """Write data to an Excel (.xlsx) file.

    Creates the file if it doesn't exist, or appends a new sheet.

    Args:
        filepath: Absolute or relative path where the .xlsx file will be saved.
        sheet_name: Name of the sheet. If omitted and file exists, uses first sheet.
        data: JSON string representing a 2D array (list of rows, each row is a list of values).
    """
    err = _check_openpyxl_installed()
    if err:
        return json.dumps({"error": err}, ensure_ascii=False)

    path = _resolve_path(filepath)
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        # Parse the data JSON
        try:
            parsed = json.loads(data)
            if not isinstance(parsed, list):
                return json.dumps({"error": "data must be a JSON array of rows (list of lists)"}, ensure_ascii=False)
        except json.JSONDecodeError as e:
            return json.dumps({"error": f"Invalid JSON data: {e}"}, ensure_ascii=False)

        # Open or create workbook
        if os.path.isfile(path):
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet to avoid conflict
            wb.remove(wb.active)

        # Determine target sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Clear existing content
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            if sheet_name:
                ws = wb.create_sheet(title=sheet_name)
            else:
                ws = wb.active
                if ws is None:
                    ws = wb.create_sheet(title="Sheet1")

        # Write data
        for row_idx, row_data in enumerate(parsed, start=1):
            if isinstance(row_data, list):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if value is not None:
                        cell.value = value

        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        wb.save(path)
        wb.close()

        return json.dumps(
            {
                "filepath": path,
                "sheet_name": ws.title,
                "rows_written": len(parsed),
                "status": "created" if not os.path.isfile(path) else "updated",
            },
            indent=2,
            ensure_ascii=False,
        )
    except Exception as exc:
        logger.exception("Failed to write Excel file: %s", exc)
        return json.dumps({"error": f"Failed to write Excel file: {exc}"}, ensure_ascii=False)


@tool("list_excel_sheets", parse_docstring=True)
def list_excel_sheets_tool(filepath: str) -> str:
    """List all sheet names in an Excel (.xlsx) workbook.

    Args:
        filepath: Absolute or relative path to the .xlsx file.
    """
    err = _check_openpyxl_installed()
    if err:
        return json.dumps({"error": err}, ensure_ascii=False)

    path = _resolve_path(filepath)
    if not os.path.isfile(path):
        return json.dumps({"error": f"File not found: {path}"}, ensure_ascii=False)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True)
        sheets = wb.sheetnames
        active = wb.active.title
        wb.close()

        return json.dumps(
            {
                "filepath": path,
                "sheets": sheets,
                "active_sheet": active,
                "total_sheets": len(sheets),
            },
            indent=2,
            ensure_ascii=False,
        )
    except Exception as exc:
        logger.exception("Failed to list Excel sheets: %s", exc)
        return json.dumps({"error": f"Failed to list Excel sheets: {exc}"}, ensure_ascii=False)
